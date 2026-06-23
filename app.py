import os
import io
import re
import base64
import json
import requests
from flask import Flask, request, jsonify, render_template, send_from_directory
from dotenv import load_dotenv, set_key
from PIL import Image as PILImage, ImageGrab
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
import shutil

# Load environment variables
load_dotenv()

app = Flask(__name__, template_folder='templates', static_folder='static')

# Never let Chrome (esp. app-mode) serve a stale cached app.js/style.css,
# otherwise UI fixes silently don't show up for the user.
@app.after_request
def add_no_cache_headers(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

@app.route('/health')
def health():
    return jsonify({"ok": True})

# Paths — all relative to this file so the app is portable across machines.
# Override the data location with the DATA_DIR env var if you want it elsewhere.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.getenv("DATA_DIR") or os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
EXCEL_PATH = os.path.join(DATA_DIR, "SAT_Performance_Tracker.xlsx")
EXCEL_TEMPLATE_PATH = os.path.join(BASE_DIR, "template", "SAT_Performance_Tracker.xlsx")
SCREENSHOTS_DIR = os.path.join(DATA_DIR, "mistakes_screenshots")

# Import dropdown categories from constants
from constants import TOPICS, QUESTION_TYPES, SUBTOPICS, ERROR_TYPES, ROOT_CAUSES, FIX_STRATEGIES
# Optional live sync to a shared Google Sheet (no-ops if unconfigured)
import gsheet_sync

# Helper: Ensure Excel file exists in Downloads
def ensure_excel_exists():
    if not os.path.exists(EXCEL_PATH):
        # Create directory if needed
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
        if os.path.exists(EXCEL_TEMPLATE_PATH):
            shutil.copy(EXCEL_TEMPLATE_PATH, EXCEL_PATH)
        else:
            raise FileNotFoundError(f"Template Excel file not found at {EXCEL_TEMPLATE_PATH}")

# Helper: Save image from base64
def save_base64_image(base64_data, output_path):
    if "," in base64_data:
        base64_data = base64_data.split(",", 1)[1]
    image_bytes = base64_data.encode('utf-8')
    image_data = base64.b64decode(image_bytes)
    
    # Open image using Pillow to verify and save
    image = PILImage.open(io.BytesIO(image_data))
    
    # Create screenshots directory
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    image.save(output_path, format="PNG")
    return output_path

# Helper: Generate compressed thumbnail for openpyxl
def create_thumbnail(image_path, row_idx):
    thumbnail_path = os.path.join(SCREENSHOTS_DIR, f"thumb_row_{row_idx}.png")
    img = PILImage.open(image_path)
    
    # Target maximum dimensions for column E (height 65 points, width 40 characters)
    # Roughly 260px width by 75px height
    max_w, max_h = 260, 75
    img.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
    img.save(thumbnail_path, format="PNG")
    return thumbnail_path

@app.route('/')
def index():
    return render_template('index.html')

def gemini_keys():
    """All configured Gemini keys, in failover order, de-duplicated."""
    keys = []
    for env in ("GEMINI_API_KEY", "GEMINI_API_KEY_2", "GEMINI_API_KEY_3", "GOOGLE_API_KEY"):
        v = (os.getenv(env) or "").strip()
        if v and v not in keys:
            keys.append(v)
    return keys

@app.route('/config', methods=['GET'])
def get_config():
    api_key = bool(gemini_keys())
    return jsonify({
        "topics": TOPICS,
        "question_types": QUESTION_TYPES,
        "subtopics": SUBTOPICS,
        "error_types": ERROR_TYPES,
        "root_causes": ROOT_CAUSES,
        "fix_strategies": FIX_STRATEGIES,
        "has_api_key": bool(api_key),
        "gsheet": gsheet_sync.status()
    })

@app.route('/settings', methods=['POST'])
def save_settings():
    data = request.json or {}
    raw = data.get("api_key", "").strip()
    if not raw:
        return jsonify({"success": False, "error": "API key cannot be empty"}), 400

    # Accept one or more keys separated by comma / space / newline.
    # They are stored as failover slots and tried in order during analysis.
    keys = [k.strip() for k in re.split(r"[\s,]+", raw) if k.strip()]

    env_path = os.path.join(BASE_DIR, ".env")

    slots = ["GEMINI_API_KEY", "GEMINI_API_KEY_2", "GEMINI_API_KEY_3"]
    for slot in slots:
        os.environ.pop(slot, None)
    for slot, key in zip(slots, keys):
        set_key(env_path, slot, key)
        os.environ[slot] = key

    n = len(keys[:len(slots)])
    return jsonify({"success": True, "message": f"Saved {n} API key{'s' if n != 1 else ''} successfully"})

@app.route('/grab-clipboard', methods=['POST'])
def grab_clipboard():
    try:
        # Pillow ImageGrab.grabclipboard() grabs from OS clipboard
        img = ImageGrab.grabclipboard()
        if isinstance(img, PILImage.Image):
            # Save to buffer and send as base64
            buffered = io.BytesIO()
            img.save(buffered, format="PNG")
            img_str = base64.b64encode(buffered.getvalue()).decode("utf-8")
            return jsonify({
                "success": True, 
                "image": f"data:image/png;base64,{img_str}"
            })
        else:
            return jsonify({
                "success": False, 
                "error": "No image found in macOS clipboard. Copy an image/screenshot first, or paste it directly using Cmd+V."
            })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

def friendly_gemini_error(response):
    """Turn a non-200 Gemini response into a short, actionable message."""
    try:
        api_err = response.json().get("error", {})
        api_msg = api_err.get("message", response.text)
        reason = ""
        for d in api_err.get("details", []):
            if d.get("reason"):
                reason = d["reason"]
                break
        status = str(api_err.get("status", ""))
    except Exception:
        api_msg, reason, status = response.text, "", ""

    if response.status_code in (401, 403) or "API_KEY" in reason or "PERMISSION_DENIED" in status:
        return ("Gemini rejected the key (HTTP "
                f"{response.status_code}{' / ' + reason if reason else ''}). "
                "The key is invalid, restricted, or the Generative Language API isn't enabled "
                "for it. Get a fresh key at https://aistudio.google.com/apikey (it starts with "
                "'AIzaSy') and save it via the gear icon. You can still fill the fields and Save manually.")
    return f"Gemini API error (HTTP {response.status_code}): {api_msg}"


@app.route('/analyze', methods=['POST'])
def analyze():
    keys = gemini_keys()
    if not keys:
        return jsonify({"success": False, "error": "Gemini API key is not configured. Please set it in the Settings panel."}), 400

    data = request.json or {}
    image_data = data.get("image")
    if not image_data:
        return jsonify({"success": False, "error": "No image data received"}), 400

    if "," in image_data:
        header, encoded_string = image_data.split(",", 1)
    else:
        encoded_string = image_data

    headers = {"Content-Type": "application/json"}

    prompt = f"""
You are an expert SAT tutor and analyzer. Analyze this screenshot of a Digital SAT question.
Extract and suggest details that fit the following fields for a student error log. 

IMPORTANT RULES FOR FIELD MATCHING:
- Section: Must be exactly "Math" or "Reading & Writing"
- Topic: Must be chosen EXACTLY from this list: {json.dumps(TOPICS)}
- Subtopic: Must be chosen EXACTLY from this list: {json.dumps(SUBTOPICS)}
- Question Type: Must be chosen EXACTLY from this list: {json.dumps(QUESTION_TYPES)}
- Error Type: Suggest the most likely choice from this list: {json.dumps(ERROR_TYPES)}
- Root Cause: Suggest the most likely choice from this list: {json.dumps(ROOT_CAUSES)}
- Fix Strategy: Suggest the most likely choice from this list: {json.dumps(FIX_STRATEGIES)}

For the other fields, autofigure them from the visual cues:
- Source / Site: Detect the source platform (e.g. "Bluebook Test 1", "Khan Academy", "UWorld", etc. from logos, colors, fonts, or headers). If not clear, default to a descriptive guess.
- Correct Answer: The correct answer (e.g. A, B, C, D or a number if it is a Math student-produced response).
- Your Answer: The answer selected by the student (look for red X markings, selected option states, or user-written inputs). If not visible, return "".
- Notes: A concise 1-sentence description of the question concept and why the student might have missed it (e.g., "Systems of linear equations with no solution, misidentified slope relationship").

Return your response strictly as a JSON object containing the following keys:
"Source / Site", "Section", "Correct Answer", "Your Answer", "Topic", "Subtopic", "Question Type", "Error Type", "Root Cause", "Fix Strategy", "Notes"

Do not wrap the output in markdown block wrappers. Return raw JSON content only.
"""

    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {
                    "inlineData": {
                        "mimeType": "image/png",
                        "data": encoded_string
                    }
                }
            ]
        }],
        "generationConfig": {
            "responseMimeType": "application/json",
            # Disable 2.5-flash "thinking" — this is extraction/classification,
            # not reasoning. Cuts analyze latency from ~13s to a few seconds.
            "thinkingConfig": {"thinkingBudget": 0}
        }
    }
    
    # Try each configured key in order; fall back to the next on any failure.
    base_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"
    last_error = "No Gemini keys available."
    for idx, api_key in enumerate(keys, 1):
        try:
            response = requests.post(f"{base_url}?key={api_key}", headers=headers, json=payload, timeout=120)
            if response.status_code == 200:
                result = response.json()
                content = result["candidates"][0]["content"]["parts"][0]["text"]
                parsed = json.loads(content.strip())
                return jsonify({"success": True, "analysis": parsed, "key_used": idx})
            last_error = friendly_gemini_error(response)
        except Exception as e:
            last_error = f"Could not reach Gemini (key {idx}): {e}"
        # otherwise loop to the next key

    prefix = f"All {len(keys)} keys failed. " if len(keys) > 1 else ""
    return jsonify({"success": False, "error": prefix + last_error}), 200

@app.route('/save', methods=['POST'])
def save_row():
    try:
        ensure_excel_exists()
        
        data = request.json or {}
        image_data = data.get("image")
        
        # Load the workbook
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Error Log" not in wb.sheetnames:
            return jsonify({"success": False, "error": "Spreadsheet does not contain 'Error Log' sheet"}), 500
            
        ws = wb["Error Log"]
        
        # Find first available row (where A is empty)
        target_row = None
        for r in range(2, 501):
            val = ws.cell(row=r, column=1).value
            if val is None or str(val).strip() == "":
                target_row = r
                break
                
        if not target_row:
            # If all 500 pre-formatted rows are filled, append a new row
            target_row = ws.max_row + 1
            
        # Set height of the target row to 65 to fit the image
        ws.row_dimensions[target_row].height = 65
        
        # Save screenshot if provided
        original_img_path = ""
        if image_data:
            original_img_path = os.path.join(SCREENSHOTS_DIR, f"mistake_row_{target_row}.png")
            save_base64_image(image_data, original_img_path)
            
            # Create compressed thumbnail and insert drawing
            thumb_path = create_thumbnail(original_img_path, target_row)
            img_drawing = OpenpyxlImage(thumb_path)
            
            # Place image in cell E
            ws.add_image(img_drawing, f"E{target_row}")
            
            # Make cell clickable by setting value and hyperlink
            cell_e = ws.cell(row=target_row, column=5)
            cell_e.value = "Mistake Clip"
            cell_e.hyperlink = f"file://{os.path.abspath(original_img_path)}"
            cell_e.font = Font(name="Segoe UI", size=9, underline="single", color="0563C1")
            cell_e.alignment = Alignment(horizontal="center", vertical="center")
            
        # Populate text values
        fields_mapping = {
            1: data.get("source_site", ""),        # A: Source / Site
            2: data.get("section", ""),            # B: Section
            3: data.get("correct_answer", ""),     # C: Correct Answer
            4: data.get("your_answer", ""),        # D: Your Answer
            # 5: E is Screenshot
            6: data.get("topic", ""),              # F: Topic
            7: data.get("subtopic", ""),           # G: Subtopic
            8: data.get("question_type", ""),      # H: Question Type
            9: data.get("error_type", ""),         # I: Error Type
            10: data.get("root_cause", ""),        # J: Root Cause
            11: data.get("fix_strategy", ""),      # K: Fix Strategy
            13: data.get("retest_status", "Not Reviewed"), # M: Retest Status
            14: data.get("notes", "")              # N: Notes
        }
        
        # Parse Time Taken
        time_val = data.get("time_taken", "")
        if time_val != "":
            try:
                fields_mapping[12] = int(time_val) # L: Time Taken (Sec)
            except ValueError:
                fields_mapping[12] = time_val
        else:
            fields_mapping[12] = ""
            
        # Populate cells & clean formatting
        no_fill = PatternFill(fill_type=None)
        font_body = Font(name="Segoe UI", size=10)
        thin_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for col_idx, val in fields_mapping.items():
            cell = ws.cell(row=target_row, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            cell.fill = no_fill  # Clear yellow fill
            
            # Alignments
            if col_idx in [3, 4, 12, 13]: # answers, time, status center
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Hidden Date Helper Formula in Column O (15th column)
        ws.cell(row=target_row, column=15, value=f'=IF(A{target_row}="", "", IFERROR(INDEX(Dashboard!$A$13:$A$35, MATCH(A{target_row}, Dashboard!$B$13:$B$35, 0)), ""))')
        ws.cell(row=target_row, column=15).number_format = "yyyy-mm-dd"
        
        # Save workbook
        wb.save(EXCEL_PATH)
        wb.close()
        
        # Best-effort live sync to the shared Google Sheet. Never breaks local
        # logging: any failure is reported but the row is already saved locally.
        sync = gsheet_sync.append_row(data, original_img_path or None)

        msg = f"Successfully logged error to row {target_row} in SAT_Performance_Tracker.xlsx"
        if sync.get("ok"):
            msg += " — also synced to your shared Google Sheet."
        elif not sync.get("skipped"):
            msg += f" (Google Sheet sync failed: {sync.get('error')})"

        return jsonify({
            "success": True,
            "message": msg,
            "row": target_row,
            "gsheet": sync
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

if __name__ == "__main__":
    ensure_excel_exists()
    # debug=False: avoid the Werkzeug reloader (double process + file-watch)
    # which is unreliable when launched detached from the .app bundle.
    port = int(os.getenv("PORT", "5001"))
    app.run(host="127.0.0.1", port=port, debug=False)
